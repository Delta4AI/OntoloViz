import io
import os
from collections import defaultdict
import sqlite3
from datetime import datetime
from difflib import get_close_matches
from re import match
from string import ascii_uppercase
from textwrap import wrap

from tkinter import Tk, messagebox
from openpyxl import Workbook, load_workbook
from openpyxl.styles import PatternFill
from plotly.graph_objects import Figure, Sunburst, Icicle, Treemap
from plotly.offline import plot as plotly_plot
from plotly.subplots import make_subplots

from src.ontoloviz.obo_utils import sanitize_string
from src.ontoloviz.core_utils import chunks, generate_color_range, prioritize_bright_colors


class SunburstBase:
    """Generalized base class"""
    def __init__(self):
        self.database = None
        self.drug_lookup = None
        self.drug_lookup_reverse = None
        self.phenotype_lookup = None
        self.phenotype_lookup_reverse = None
        self.zero = 0.000001337
        self.fake_one = 1.000001337
        self.thread_status = ""
        self.thread_return = None
        self.custom_ontology = None
        self.custom_ontology_title = None
        self.plot_error = None

        # settings
        self.s = None
        self.init_settings()

    def init_db(self, database: str = None):
        """Initializes database and creates lookup dictionaries

        :param database: path to database"""
        self.database = database

        # entity lookup tables
        self.drug_lookup = {_[0]: _[1] for _ in self.query("SELECT drug_name, id FROM drug_lookup")}
        self.drug_lookup_reverse = {v: k for k, v in self.drug_lookup.items()}
        self.phenotype_lookup = {_[0]: _[1] for _ in self.query("SELECT phenotype_name, id "
                                                                "FROM phenotype_lookup")}
        self.phenotype_lookup_reverse = {v: k for k, v in self.phenotype_lookup.items()}

    def verify_db(self, fn: str = None) -> bool:
        """Verifies database integrity by checking available tables

        :returns: True if database contains all required columns
        """
        ret = self.query("SELECT name FROM sqlite_master "
                         "WHERE type='table' AND name NOT LIKE 'sqlite_%'",
                         database=fn)
        required_columns = ['pheno_indirect_semantic', 'pheno_indirect_explicit',
                            'pheno_direct_explicit', 'pheno_direct_semantic', 'drug_atc',
                            'phenotype_lookup', 'drug_lt', 'mesh_tree', 'drug_lookup']
        if set([_[0] for _ in ret]) == set(required_columns):
            print("Database verified")
            return True
        return False

    @staticmethod
    def verify_file(fn: str = None) -> [None, str]:
        """Verifies a MeSH-Tree Excel/.tsv file

         - .tsv files are verified based on number of columns in first tab
         - .xlsx files are verified based on mesh_file or atc_excel = True in tab 'Settings'
           and number of columns in tab 'Tree'

        :param fn: Path to MeSH-tree Excel file
        :returns: "mesh_excel" or "atc_excel" if file is valid
        """
        if os.path.splitext(fn)[-1] == ".tsv":
            with open(fn, mode="r", encoding="utf-8") as f:
                for idx, line in enumerate(f):
                    if idx == 0:
                        columns = line.rstrip("\n").split("\t")
                        if len(columns) == 6 and columns[0] == "ATC code":
                            file_type = "atc_tsv"
                        elif len(columns) == 7:
                            file_type = "mesh_tsv"
                        else:
                            return None
                            # raise ValueError("TSV verification failed. Expected 6 columns for "
                            #                  "ATC-tree, 7 columns for MeSH tree")
                        print(f"TSV verified as '{file_type}': {fn}")
                        return file_type

        else:
            workbook = load_workbook(fn, read_only=True)

            # key equals 'Settings' tab in excel; value = number of columns to verify in 'Tree' tab
            req = {"mesh_excel": 7, "atc_excel": 6}

            # check for mesh_excel/atc_excel = True in 'Settings' tab
            try:
                file_type = [r[0].value for r in workbook["Settings"].rows
                             if r[0].value in req.keys() and r[1].value == True]
                if not file_type:
                    raise KeyError("Excel verification failed: no valid Setting"
                                   " for keys 'mesh_excel' or 'atc_excel' found.")
            # in case tab 'Settings' does not exist
            except KeyError as exc:
                # classify solely based on column number in first sheet
                cols = workbook.worksheets[0].max_column
                if cols not in req.values():
                    return None
                    # raise ValueError(
                    #     "Excel verification without settings failed: Amount of columns does not "
                    #     "match any known configuration!\nThis files columns: "
                    #     f"{cols}\nPossible values: {req}\nException: {exc}") from exc
                flipped_req = {v: k for k, v in req.items()}
                return flipped_req[cols] + "_no_settings"

            # check for number of columns in 'Tree' tab
            if workbook["Tree"].max_column != req[file_type[0]]:
                return None
                # raise ValueError("Excel verification failed: Columns in tab 'Tree' do not match "
                #                  "expected number. Expected: "
                #                  f"{req[file_type[0]]}, actual: {workbook['Tree'].max_column}")

            print(f"Excel verified as '{file_type[0]}': {fn}")
            return file_type[0]

    def set_thread_status(self, text):
        """Sets thread status and prints text"""
        self.thread_status = text
        if text:
            print(text)

    def set_color_scale(self, color_scale: list = None):
        """Sets color scale and default color

        :param color_scale: color-scale in format [[0, '#FFFFFF'], [1, '#FF0000']]
            where list[0] must be in range 0..1 and amount of lists can vary
        """
        self.set_settings({"color_scale": color_scale, "default_color": color_scale[0][1]})

    def get_label_to_current_counts(self, current_data: list) -> dict:
        """
        Takes current_data lists and populates a list with dictionaries in the current trees
        structure to map labels to the currently displayed counts
        :param current_data: list of tuples containing custom data used for plotting
        :return: dictionary with label as key and current count as value
        """
        return {node[0]: node[1] for sub_tree in current_data for node in sub_tree}

    def set_settings(self, settings: dict = None):
        """Verifies/converts settings, example call: self.set_settings({'show_border': 'True'})"""
        for _k, _v in settings.items():
            if _k not in self.s.keys():
                raise KeyError(f"Illegal settings key used: '{_k}'")

            # resolve booleans
            if _k in ["show_border", "export_plot", "mesh_drop_empty_last_child",
                     "atc_propagate_enable", "mesh_propagate_enable"]:
                if _v in ["True", "TRUE", "1", 1]:
                    _v = True
                elif _v in ["False", "FALSE", "0", 0]:
                    _v = False
                else:
                    raise ValueError(f"Illegal value for setting '{_k}': '{_v}' - boolean required")

            # resolve ints
            if _k in ["atc_propagate_lvl", "mesh_propagate_lvl"]:
                try:
                    _v = int(_v)
                except ValueError:
                    raise ValueError(f"Illegal value for setting '{_k}': '{_v}' - integer required")

            # resolve floats
            if _k in ["border_width"]:
                try:
                    _v = float(_v)
                except ValueError:
                    raise ValueError(
                        f"Illegal value for setting '{_k}': '{_v}' - integer or float required")

            # custom resolves
            if _k == "atc_labels" and _v not in ["all", "propagation", "drugs", "none"]:
                raise ValueError(f"Illegal value for setting '{_k}': '{_v}' - "
                                 f"valid are 'all', 'propagation', 'drugs', 'none'")

            if _k == "mesh_labels" and _v not in ["all", "propagation", "none"]:
                raise ValueError(f"Illegal value for setting '{_k}': '{_v}' "
                                 "- valid are 'all', 'propagation', 'none'")

            if _k == "atc_wedge_width" and _v not in ["total", "remainder"]:
                raise ValueError(
                    f"Illegal value for setting '{_k}': '{_v}' - valid are 'total', 'remainder'")

            if _k == "default_color" and not match("#[a-fA-F0-9]{6}$", _v):
                raise ValueError(
                    f"Illegal value for setting '{_k}': '{_v}' - valid format is '#FFFFFF'")

            if _k in ["atc", "mesh_summary_plot"] and (_v < 0 or _v > 20):
                raise ValueError(
                    f"Illegal value for setting '{_k}': '{_v}' - valid are integers > 0 and < 20")

            if _k == "atc_propagate_color" and _v not in ["specific", "global", "off"]:
                raise ValueError(f"Illegal value for setting '{_k}': '{_v}' "
                                 "- valid are 'specific', 'global' and 'off'")

            if _k == "mesh_propagate_color" and _v not in ["specific", "global",
                                                           "off", "phenotype"]:
                raise ValueError(f"Illegal value for setting '{_k}': '{_v}' "
                                 "- valid are 'specific', 'global', 'phenotype' and 'off'")

            if _k in ["atc_propagate_counts", "mesh_propagate_counts"] \
                    and _v not in ["off", "level", "all"]:
                raise ValueError(f"Illegal value for setting '{_k}': '{_v}' "
                                 "- valid are 'off', 'level' and 'all'")

            # apply setting
            self.s[_k] = _v
            print(f"Loaded setting: {_k} - {_v}")

    def init_settings(self):
        """Initializes settings with default values"""
        self.s = {
            "color_scale": [[0, "#FFFFFF"], [0.2, "#403C53"], [1, "#C33D35"]],
            "show_border": True,
            "border_color": "rgba(0,0,0,0.25)",
            "border_width": 1,
            "export_plot": False,

            # relevant only for MeSH data
            "mesh_drop_empty_last_child": False,
            "mesh_propagate_enable": False,
            "mesh_propagate_lvl": 0,
            "mesh_propagate_color": "specific",
            "mesh_propagate_counts": "off",
            "mesh_labels": "all",
            "mesh_summary_plot": 5,  # 0 equals disabled
            "mesh_excel": True,

            # relevant only for ATC data
            "atc_propagate_enable": False,
            "atc_propagate_lvl": 5,
            "atc_propagate_color": "specific",
            "atc_propagate_counts": "off",
            "atc_labels": "all",
            "atc_wedge_width": "total",
            "atc_summary_plot": 5,  # 0 equals disabled
            "atc_excel": True,
        }

        self.s["default_color"] = self.s["color_scale"][0][1]

    def query(self, query: str = None, query_args: list = None, database: str = None) -> list:
        """Execute query, fetch and return all results

        :param query: SQL query - column and table names have to be string-formatted,
            WHERE arguments handed over as list
        :param query_args: optional arguments for query
        :param database: Optional path to database, required for db-verification step
        :returns: List of tuples

        Example call:
          ret = query("test.db", "SELECT {} FROM {} WHERE {}=?"
                .format("col1", "test_table", "col2"), ["value1"])
        """
        if not query_args:
            query_args = []
        if not database:
            database = self.database
        with sqlite3.connect(database) as connection:
            cursor = connection.cursor()
            cursor.execute(query, query_args)
            ret = cursor.fetchall()
            cursor.close()
            return ret

    def get_entity_id(self, entity_name: str = None, entity_type: str = None) -> int:
        """Get entity id from lookup dict (case-insensitive)

        :param entity_name: Entity name to be resolved to id
        :param entity_type: Type of entity, must be in ["drug", "phenotype"]
        :raises ValueError: Error message including alternative drug names
        :returns: ID of given entity
        """
        lookup, name = None, None
        if entity_type == "drug":
            lookup = self.drug_lookup
            name = entity_name.upper()
        elif entity_type == "phenotype":
            lookup = self.phenotype_lookup
            name = entity_name.title()
        matches = get_close_matches(name, [str(_) for _ in lookup.keys()])
        if matches and name == matches[0]:
            return lookup[matches[0]]
        elif matches:
            raise ValueError(f"Could not resolve {entity_type} '{entity_name}', "
                             f"did you mean: {matches} ?")
        else:
            raise ValueError(f"Could not resolve {entity_type} '{entity_name}'")

    def get_phenotype_id(self, phenotype_name: str = None) -> int:
        """Get phenotype id from lookup dict

        :param phenotype_name: Phenotype name, e.g. 'Headache'
        :returns: ID of phenotype name
        """
        return self.get_entity_id(phenotype_name, "phenotype")

    def get_drug_id(self, drug_name: str = None) -> int:
        """Get drug id from lookup dict

        :param drug_name: Drug name, e.g. 'Aspirin'
        :returns: ID of drug name
        """
        return self.get_entity_id(drug_name, "drug")

    def calculate_color_scale_for_node(self, sub_tree: dict = None, max_val: int = None,
                                       max_level: [str, int] = None,
                                       count_key: str = "counts") -> tuple:
        """Get color scale based on max value of counts of children in a subtree

        :param sub_tree: MeSH/ATC subtree dictionary (e.g. mesh_tree['C01']) (e.g. atc_tree['L'])
        :param max_val: calculates color scale based on given value
        :param max_level: calculates max_val only up to given level, or up to last level
            when 'last' is given
        :param count_key: key from node to consider for max counts
            (ATC sunburst requires displayed_counts)
        :returns: tuple, where first index is factor, second index is List of hex colors for all
            available counts, where index = amount of counts, in case factorization is True, amount
            of colors are divided by 10 to speed up code
        """
        factor = 1
        try:
            if not max_val:

                # calculate maxima for entire sub tree
                if not max_level:
                    max_val = max([_[count_key] for _ in sub_tree.values()])
                else:

                    # calculate maxima based on level
                    if isinstance(max_level, int):
                        max_val = max([_[count_key] for _ in sub_tree.values()
                                       if _["level"] >= max_level])

                    # calculate maxima based on most outer nodes
                    # only works for keys with C.C (dot) annotation
                    elif isinstance(max_level, str):
                        whitelist = set()
                        max_val = 0
                        for k, v in sub_tree.items():
                            if k not in whitelist:
                                if v[count_key] >= max_val:
                                    max_val = v[count_key]

                            # add all parents to whitelist
                            for idx in range(k.count(".")):
                                whitelist.add(k.rsplit(".", idx+1)[0])

                # convert to int
                max_val = int(max_val)

            # factor calculation
            if 100000 <= max_val < 250000:
                factor = 10
                max_val = max_val / 10
            elif max_val >= 250000:
                factor = 25
                max_val = max_val / 25
        except ValueError:
            max_val = 0

        scale = [self.s["default_color"]]  # create list with default color as first item
        for i in range(len(self.s["color_scale"]) - 1):
            lower_limit, lower_color = self.s["color_scale"][i]
            upper_limit, upper_color = self.s["color_scale"][i + 1]
            low_cutoff = int(max_val * lower_limit)
            high_cutoff = int(max_val * upper_limit)
            scale.extend(generate_color_range(lower_color, upper_color, high_cutoff - low_cutoff))

        self.set_thread_status(f"Generating color scale for {int(max_val)} "
                               f"(factor: {factor}) values ..")
        return factor, scale

    def get_total_counts(self, count_key: str = "counts") -> float:
        """Sums up counts of tree

        :param count_key: key in children elements that contains value to weight
        :returns: sum of all values of children's count_key
        """
        if isinstance(self, ATCSunburst):
            return sum([int(vv[count_key]) for k, v in self.atc_tree.items()
                        for kk, vv in v.items()])
        elif isinstance(self, MeSHSunburst):
            return sum([int(vv[count_key]) for k, v in self.mesh_tree.items()
                        for kk, vv in v.items()])

    def export_settings(self, fn: [str, None] = None, wb: Workbook = None,
                        settings: list = None) -> str:
        """Subroutine to write settings to workbook

        :param fn: target filename
        :param wb: target workbook
        :param settings: settings to write/update
        :returns: absolute path of created Excel file
        """
        if not wb:
            wb = load_workbook(fn)

        ws = wb.get_sheet_by_name("Settings")

        # set column widths based on max cell size without much logic
        ws.column_dimensions["A"].width = len(max(settings, key=lambda x: len(str(x[0])))[0]) + 2
        ws.column_dimensions["B"].width = len(max(settings, key=lambda x: len(str(x[1])))[1]) + 2

        # write rows to worksheet
        for s in settings:
            try:
                ws.append(s)
            except ValueError:
                # for [[0, '#FFFFFF'], [0.2, '#403C53'], [1, '#C33D35']]
                ws.append(tuple(str(_) for _ in s))

        return self.save_workbook(fn=fn, wb=wb)

    @staticmethod
    def save_workbook(fn: str = None, wb: Workbook = None) -> str:
        """Save workbook under given filename; If file exists, appends timestamp to filename

        :param fn: excel filename
        :param wb: Workbook object
        :returns: absolute path of created Excel file
        """
        try:
            wb.save(fn)
        except PermissionError:
            print("\tFile already exists - appending timestamp ..")
            fn = os.path.splitext(fn)[0] + f"_{datetime.now().strftime('%Y-%m-%d_%H-%M-%S')}.xlsx"
            wb.save(fn)
        finally:
            wb.close()
        print(f"\tExported to: {os.path.abspath(fn)}")
        return os.path.abspath(fn)

    def export_tree_to_excel(self, fn: str = None, header: list = None, rows: list = None,
                             settings: list = None, color_col: int = None) -> str:
        """Export Tree to Excel

        :param fn: target filename
        :param header: header row as list of strings
        :param rows: data rows in tab 'Tree' as list of tuples
        :param settings: settings rows in tab 'Settings' as list of tuples
        :param color_col: index (starts with 1) of column to colorize in 'Tree' tab
            based on hex-code in cell

        :returns: absolute path of exported Excel file
        """
        # create workbook, write header
        wb = Workbook()
        ws = wb.active
        ws.title = "Tree"
        ws.append(header)

        # set column widths based maxvalue of the cells
        # (lower-limit = header length, upper-limit = 100)
        col_width = [len(_) for _ in header]
        for idx, (width, col_letter) in enumerate(zip(col_width, ascii_uppercase[:len(col_width)])):
            max_row_width = len(str(max(rows, key=lambda x: len(str(x[idx])))[idx]))
            current_col_width = int(width)
            if width <= max_row_width <= 100:
                current_col_width = max_row_width
            elif max_row_width > 100:
                current_col_width = 100
            ws.column_dimensions[col_letter].width = current_col_width + 2

        # write rows to worksheet, apply color to cells
        for row in rows:
            ws.append(row)
            fg_col = row[color_col-1].lstrip("#")
            ws.cell(row=ws.max_row, column=color_col).fill = PatternFill(fgColor=fg_col,
                                                                         fill_type="solid")

        # append settings to second tab, write & close workbook
        wb.create_sheet(title="Settings", index=1)
        return self.export_settings(fn=fn, wb=wb, settings=settings)

    @staticmethod
    def export_tree_to_tsv(fn: str = None, header: list = None, rows: list = None) -> str:
        """Export Tree to TSV

        :param fn: target filename
        :param header: header row as list of strings
        :param rows: data rows in tab 'Tree' as list of tuples

        :returns: absolute path of exported Excel file
        """

        out_file = None
        try:
            out_file = open(fn, mode="w", encoding="utf-8")
        except PermissionError:
            print("\tFile already exists - appending timestamp ..")
            fn = os.path.splitext(fn)[0] + f"_{datetime.now().strftime('%Y-%m-%d_%H-%M-%S')}.tsv"
            out_file = open(fn, mode="w", encoding="utf-8")
        finally:
            # write header
            out_file.write("\t".join(header) + "\n")

            # write lines
            for row in rows:
                out_file.write("\t".join([str(_) for _ in row]) + "\n")

            out_file.close()

        print(f"\tExported to: {os.path.abspath(fn)}")
        return os.path.abspath(fn)

    def tree_color_propagation(self, plot_tree: dict = None, count_key: str = None):
        """Apply color propagation to a tree based on current settings

        :param plot_tree: tree as dictionary
        :param count_key: must be in ['counts', 'imported_counts']
        """
        propagation_enabled, propagation_type, max_level = None, None, None
        if isinstance(self, MeSHSunburst):
            propagation_enabled = self.s["mesh_propagate_enable"]
            propagation_type = self.s["mesh_propagate_color"]
            max_level = self.s["mesh_propagate_lvl"]
        elif isinstance(self, ATCSunburst):
            propagation_enabled = self.s["atc_propagate_enable"]
            propagation_type = self.s["atc_propagate_color"]
            max_level = self.s["atc_propagate_lvl"]

        scale, specific_scales, factor = None, [], None
        if propagation_enabled and propagation_type in ["specific", "global", "phenotype"]:
            self.set_thread_status(f"Propagating color ({propagation_type}) ..")

            # calculate global color scale on type 'global'
            if propagation_type == "global":
                max_counts = max(_["imported_counts"] for d in plot_tree.values()
                                 for _ in d.values() if _["level"] >= max_level)
                factor, scale = self.calculate_color_scale_for_node(max_val=max_counts)

            # calculate individual color scales for each sub tree on type 'specific'
            elif propagation_type == "specific":
                for k, v in plot_tree.items():
                    factor, scale = self.calculate_color_scale_for_node(v, max_level=max_level,
                                                                        count_key=count_key)
                    specific_scales.append((factor, scale))

            # calculate individual color scales based on the most outer phenotypes
            # of each sub-tree on type 'phenotype'
            elif propagation_type == "phenotype":
                for k, v in plot_tree.items():
                    factor, scale = self.calculate_color_scale_for_node(v, max_level="last",
                                                                        count_key=count_key)
                    specific_scales.append((factor, scale))

            # iterate over nodes, apply color if level is in accepted range
            for idx, (k, v) in enumerate(plot_tree.items()):

                # empty whitelist for 'phenotype' color propagation
                whitelist = set()

                # get specific scales
                if propagation_type in ["specific", "phenotype"]:
                    factor, scale = specific_scales[idx]

                # iterate over nodes
                for kk, vv in v.items():

                    # for 'phenotype', check if vv["level"] is the most outer node and apply color
                    if propagation_type == "phenotype":
                        if kk not in whitelist:

                            # add all parents to whitelist
                            for dot_idx in range(kk.count(".")):
                                whitelist.add(kk.rsplit(".", dot_idx+1)[0])

                            # apply color
                            plot_tree[k][kk]["color"] = scale[int(vv["imported_counts"]/factor)]
                        else:
                            plot_tree[k][kk]["color"] = self.s["default_color"]

                    # for other types, apply based on level
                    else:
                        if vv["level"] >= max_level:
                            plot_tree[k][kk]["color"] = scale[int(vv["imported_counts"] / factor)]

                        else:
                            plot_tree[k][kk]["color"] = self.s["default_color"]

    def generate_plot_supplements(self, plot_tree: dict = None) -> tuple:
        """Generates nested lists for subtrees containing label, percentage, custom data;
         creates filtered plot tree based on drop empty setting

        :param plot_tree: dictionary containing trees and nodes
        :return: tuple of lists containing labels and percentages for each node in each subtree
        """
        label_mode, propagate_count_mode, propagate_lvl, hover_template = None, None, None, None
        propagate_color_mode = None
        propagate_enabled = None
        specific_color_propagation = False

        if isinstance(self, MeSHSunburst):
            label_mode = self.s["mesh_labels"]
            propagate_count_mode = self.s["mesh_propagate_counts"]
            propagate_color_mode = self.s["mesh_propagate_color"]
            propagate_lvl = self.s["mesh_propagate_lvl"]
            hover_template = ("%{customdata[0]}: <b>%{customdata[1]}</b> (%{customdata[2]}%)"
                              "<br>--<br>"
                              "Label: %{customdata[3]}"
                              "<br>"
                              "Tree ID: %{customdata[4]}"
                              "<br>"
                              "Children: %{customdata[5]}"
                              "<br>--<br>"
                              "%{customdata[6]}"
                              "%{customdata[7]}"
                              "<extra></extra>")
            propagate_enabled = self.s["mesh_propagate_enable"]
        elif isinstance(self, ATCSunburst):
            label_mode = self.s["atc_labels"]
            propagate_count_mode = self.s["atc_propagate_counts"]
            propagate_color_mode = self.s["atc_propagate_color"]
            propagate_lvl = self.s["atc_propagate_lvl"]
            hover_template = ("%{customdata[0]}: <b>%{customdata[1]}</b> (%{customdata[2]}%)"
                              "<br>--<br>"
                              "ATC code: %{customdata[3]}"
                              "<br>"
                              "Children: %{customdata[4]}"
                              "%{customdata[5]}"
                              "<extra></extra>")
            propagate_enabled = self.s["atc_propagate_enable"]

        if propagate_enabled and propagate_color_mode == "specific":
            specific_color_propagation = True

        # get max counts to adapt percentages in case global colors are used
        global_sum = int(sum([max(_["imported_counts"] for _ in sub.values())
                              for sub in plot_tree.values()]))

        # populate labels and percentages
        custom_ontology_counts = None
        if self.custom_ontology:
            custom_ontology_counts = self._get_child_sums(plot_tree)
        labels, custom_data = [], []
        for idx, (k, v) in enumerate(plot_tree.items()):
            self.thread_status = f"Creating plot supplements .. {idx}/{len(plot_tree)}"
            wedge_labels, custom_tuples, node_percentage = [], [], None
            sub_tree_sum = int(sum(x["imported_counts"] for x in plot_tree[k].values()))
            propagate_threshold_sum = int(sum(x["imported_counts"] for x in plot_tree[k].values()
                                              if x["level"] >= propagate_lvl))

            if not sub_tree_sum:
                sub_tree_sum = 1
            if not propagate_threshold_sum:
                propagate_threshold_sum = 1

            for kk, vv in v.items():

                # wedge labels
                wrapped_label = "<br>".join(wrap(vv.get("label", ""), 20))
                if label_mode == "all":
                    wedge_labels.append(wrapped_label)
                elif label_mode == "propagation":
                    wedge_labels.append(wrapped_label if vv["level"] >= propagate_lvl else "")
                elif label_mode == "drugs":
                    wedge_labels.append(wrapped_label if vv["level"] == 5 else "")
                elif label_mode == "none":
                    wedge_labels.append("")

                # percentages
                try:
                    if propagate_enabled and propagate_color_mode == "global":
                        node_percentage = round(vv["imported_counts"] / global_sum * 100, 1)
                    else:
                        if propagate_count_mode in ["off", "all"] or not propagate_enabled:
                            node_percentage = round(vv["imported_counts"] / sub_tree_sum * 100, 1)
                        elif propagate_count_mode == "level":
                            if vv["level"] >= propagate_lvl:
                                node_percentage = round(
                                    vv["imported_counts"] / propagate_threshold_sum * 100)
                            else:
                                node_percentage = round(
                                    vv["imported_counts"] / sub_tree_sum * 100, 1)
                except ZeroDivisionError:
                    node_percentage = 0

                # custom data
                hover_label = vv.get("label", "Undefined")
                count = int(vv["imported_counts"])
                node_id = vv["id"]
                if custom_ontology_counts:
                    child_sum = custom_ontology_counts[k][kk]
                else:
                    child_sum = sum(
                        [1 for z in v.keys() if z.startswith(vv["id"]) and z != vv["id"]])
                comment = str("<br>--<br>" + "<br>".join(wrap("Comment: " + vv["comment"], 65))
                              if vv.get("comment", None) else "")

                if isinstance(self, MeSHSunburst):
                    custom_tuples.append(
                        (hover_label, count, node_percentage, vv.get("mesh_id", hover_label),
                         node_id, child_sum,
                         "<br>".join(wrap("Description: " + vv["description"], 65)), comment))
                elif isinstance(self, ATCSunburst):
                    custom_tuples.append(
                        (hover_label, count, node_percentage, node_id, child_sum, comment))

            custom_data.append(custom_tuples)
            labels.append(wedge_labels)

        return labels, custom_data, hover_template, specific_color_propagation

    def _get_child_sums(self, plot_tree: dict = None) -> dict:
        """Creates dictionary with total amount of children for each node in each sub-tree

        :param plot_tree: dictionary with plot-ready ontology
        :return: dictionary with same structure and amount of total children for each node id
        """
        sum_dict = {}
        for sub_tree_id, sub_tree in plot_tree.items():
            sum_dict[sub_tree_id] = defaultdict(int)
            for node_id, node in sub_tree.items():

                # for root-node, use total sum for subtree and continue
                if node_id == sub_tree_id:
                    sum_dict[sub_tree_id][node_id] = len(sub_tree)
                    continue

                # increment count for direct parent
                parent = node["parent"]
                sum_dict[sub_tree_id][parent] += 1

                # traverse parents up until root-node and increment counters
                # parents = node["parent"].split("|")
                # for parent in parents:
                #     self._update_parent_counts(sub_tree_id, parent, sum_dict, sub_tree)
                while True:
                    parent = sub_tree[parent]["parent"]
                    if not parent:
                        break
                    sum_dict[sub_tree_id][parent] += 1

        return sum_dict

    # @staticmethod
    # def _update_parent_counts(sub_tree_id: str, parent: str = None,
    #                           sum_dict: dict = None, sub_tree: dict = None) -> None:
    #     """Traverse parents up"""
    #     sum_dict[sub_tree_id][parent] += 1
    #
    #     while True:
    #         parent_ids = parent.split("|")
    #         parent = None
    #         for parent_id in parent_ids:
    #             if parent_id in sub_tree:
    #                 parent = sub_tree[parent_id]["parent"]
    #                 sum_dict[sub_tree_id][parent_id] += 1
    #
    #         if parent is None:
    #             break

    def _add_color_scale_to_trace(self, trace: Sunburst, cmax: int = None,
                                  cmap: list = None) -> None:
        """Adds a color scale (legend) to a trace"""
        if not cmap:
            cmap = self.s["color_scale"]
        trace.marker.colorscale = cmap
        trace.marker.cmin = 0
        if cmax == 0:
            cmax = 1
        trace.marker.cmax = cmax
        trace.marker.colorbar = {"title": "values"}

    def _set_default_row_data(self, entity_id: str = None, label: str = None,
                              description: str = None, counts: str = None,
                              color: str = None) -> tuple:
        """Converts row data and sets default if cells are empty"""

        # set defaults if cell is empty
        if not color or not match("#[a-fA-F0-9]{6}$", color) or color == "":
            color = self.s["default_color"]

        # required .tsv conversions
        if isinstance(counts, str):
            if not counts:
                counts = 0
            counts = int(counts)

        # set zero-counts to arbitrary low number to ensure display (value must be >0)
        # if cell is empty, set to 0
        if counts == 0 or counts == 0.0 or counts is None or counts == "":
            counts = self.zero  # rounded to 0 in plot

        if not description:
            description = ""

        if not label or label == "":
            label = entity_id

        return label, description, counts, color

    def create_sunburst_figure(self, plot_tree: dict = None):
        """Create list of sunburst traces
        TODO: progressive rendering with on_click events to improve performance of large ontologies

        :param plot_tree: plot tree as dict
        """
        self.set_thread_status("Creating traces ..")

        # create list of labels, percentages
        (labels, custom_data, hover_template,
         specific_color_propagation) = self.generate_plot_supplements(plot_tree=plot_tree)
        counts_max = [max([_[1] for _ in c_data]) for c_data in custom_data]

        weighted_scale = []
        global_scale = {}
        for sub_tree, max_count in zip(plot_tree.values(), counts_max):
            sub_scale = []
            for node in sub_tree.values():
                if node["color"]:
                    global_scale[node["imported_counts"]] = node["color"]
                try:
                    val = round(node["imported_counts"]/max_count, 3)
                    if val > 1.0:
                        continue
                    sub_scale.append((val, node["color"]))
                except ZeroDivisionError:
                    continue
            if not sub_scale or len(sub_scale) < 2:
                sub_scale = [(0.0, "#FFFFFF"), (1.0, "#FFFFFF")]
            weighted_scale.append(sorted(list(set(prioritize_bright_colors(sub_scale)))))

        global_scale = sorted(global_scale.items())
        global_scale = [(round(idx/max(global_scale)[0], 3), col) for (idx, col) in global_scale]
        global_scale = prioritize_bright_colors(global_scale)

        if isinstance(self, MeSHSunburst):
            plot_type = self.plot_type.get(self.s.get("plot_type", None), Sunburst)
        else:
            plot_type = Sunburst

        # create list of traces
        traces = [plot_type(
            labels=labels[idx],
            parents=[_["parent"] for _ in v.values()],
            values=[_["counts"] for _ in v.values()],
            ids=[_["id"] for _ in v.values()],
            branchvalues=str("remainder" if isinstance(self, MeSHSunburst)
                             else self.s["atc_wedge_width"]),
            customdata=custom_data[idx],
            hovertemplate=hover_template,
            marker={'colors': [_["color"] for _ in v.values()],
                    'line': {'color': self.s["border_color"],
                             'width': self.s["border_width"]} if self.s["show_border"] else None}
        ) for idx, v in enumerate(plot_tree.values())]

        # plot configuration
        config = {"displaylogo": False,
                  "responsive": False,
                  "scrollZoom": True,
                  "displayModeBar": True,
                  "showLink": False,
                  "toImageButtonOptions": {
                      "format": "png",  # one of png, svg, jpeg, webp
                      # download at the currently-rendered size by setting height and width to None
                      "height": None,
                      "width": None,
                      "scale": 3  # Multiply title/legend/axis/canvas sizes by this factor
                  }}

        # generate headers
        headers, summary_plot, title, file_name = None, None, None, None
        if isinstance(self, MeSHSunburst):
            headers = [v[k]["label"] for k, v in sorted(self.mesh_tree.items())
                       if k in plot_tree.keys()]
            summary_plot = self.s["mesh_summary_plot"]
            if self.custom_ontology:
                title = f"{self.custom_ontology_title} Sunburst"
            else:
                title = "Phenotype Sunburst"
            title += ["", " Overview"][bool(summary_plot)]
            if self.drug_name:
                title += f" for {self.drug_name}"
                file_name = f"phenotype_sunburst_{self.drug_name.lower().replace(' ', '_')}.html"
            else:
                file_name = f"custom_sunburst_{datetime.now().strftime('%Y%M%d')}.html"
        elif isinstance(self, ATCSunburst):
            headers = [f"{k}: {v[k]['label'].title()}" for k, v in sorted(self.atc_tree.items())
                       if k in plot_tree.keys()]
            summary_plot = self.s["atc_summary_plot"]
            title = str("Drug Sunburst" + ["", " Overview"][bool(summary_plot)]
                        + f" for {self.phenotype_name}")
            file_name = f"drug_sunburst_{self.phenotype_name.lower().replace(' ', '_')}.html"

        # traces[0].marker["colorscale": self.s["color_scale"], "cmin": 0, "cmax": 100,
        # "colorbar": {"title": "values"}]

        # create figure
        self.set_thread_status("Creating figure ..")
        if summary_plot != 0:

            # add color-bar to first trace based on maximum counts; disabled for summary plots
            # with specific color propagation, as each plot would require an individual scale
            if not specific_color_propagation and self.s.get("legend", None):
                if len(global_scale) > 1:
                    self._add_color_scale_to_trace(trace=traces[0], cmax=max(counts_max),
                                                   cmap=global_scale)
                else:
                    self._add_color_scale_to_trace(
                        trace=traces[0], cmax=max(counts_max),
                        cmap=weighted_scale[counts_max.index(max(counts_max))])

            # figure for overview plot
            fig = self.generate_subplot_figure(cols=summary_plot, traces=traces, headers=headers,
                                               title=title)
        else:
            if self.s.get("legend", None):
                # color-bar for each trace
                for trace, max_count, cmap in zip(traces, counts_max, weighted_scale):
                    self._add_color_scale_to_trace(trace=trace, cmax=max_count, cmap=cmap)

            # figure for specific plots - create buttons
            buttons = []
            for i in range(len(traces)):
                specific_title = None
                if isinstance(self, MeSHSunburst):
                    specific_title = f"Counts for term {headers[i]}"
                    if self.drug_name:
                        specific_title += f" and {self.drug_name}"
                elif isinstance(self, ATCSunburst):
                    specific_title = f"Counts for term {headers[i].split(':')[-1].title()}"
                    if self.phenotype_name:
                        specific_title += f" and {self.phenotype_name}"
                buttons.append({"label": headers[i],
                                "method": "update",
                                "args": [{"visible": [i == j for j in range(len(traces))]},
                                         {"title": specific_title}]})

            # create menu and layout
            menu = [{"active": -1,
                     "buttons": buttons,
                     "yanchor": "bottom",
                     "pad": {"t": 2, "b": 10},
                     "x": 0.5,
                     "xanchor": "center"}]
            layout = {"title": {"text": title,
                                "x": 0.5,
                                "xanchor": "center"},
                      "showlegend": False,
                      "updatemenus": menu}

            # create figure, hide initial data
            fig = Figure(data=traces, layout=layout)
            fig.update_traces(visible="legendonly")

        # save / plot figure
        if self.s["export_plot"]:
            # fig.update_layout(legend=dict(x=0, y=1), autosize=False, width=1280, height=900)
            plotly_plot(fig, config=config, filename=file_name)
            html_path = os.path.abspath(file_name)
            tsv_path = None
            if isinstance(self, MeSHSunburst):
                # TODO: fix proper display when using custom_data, adapt ATC accordingly:
                # self.export_mesh_tree(mode="TSV", template=False, current_data=custom_data)
                tsv_path = self.export_mesh_tree(mode="TSV", template=False)
            elif isinstance(self, ATCSunburst):
                tsv_path = self.export_atc_tree(mode="TSV", template=False)
            self.set_thread_status(f"Exported plot to: {html_path}")
            self.thread_return = (html_path, tsv_path)

            # export template as is currently configured

        else:
            self.set_thread_status("Sunburst created")
            fig.show(config=config)

    @staticmethod
    def generate_subplot_figure(cols: int = None, traces: list = None,
                                headers: list = None, title: str = None) -> Figure:
        """Generates a subplot figure based on list of Sunburst traces and settings

        :param cols: Number of columns
        :param traces: List of Sunburst plotly objects
        :param headers: list of strings containing subheaders
        :param title: overall plot title
        """
        # helper dictionary to convert index to column/row
        idx_to_grid = {}
        for col_idx, row_idxs in enumerate(chunks(input_list=list(range(len(headers))),
                                                  number_of_chunks=cols)):
            for row_idx, original_idx in enumerate(row_idxs):
                idx_to_grid[original_idx] = (col_idx, row_idx)

        # # update domain of traces to reflect grid structure
        # for idx in range(len(traces)):
        #     traces[idx].domain = dict(column=idx_to_grid[idx][0], row=idx_to_grid[idx][1])

        fig = make_subplots(rows=max(idx_to_grid.values(), key=lambda x: x[1])[-1] + 1,
                            cols=cols,
                            specs=[[{"type": "sunburst"}
                                    for row in range(cols)]
                                   for col in range(max(idx_to_grid.values(),
                                                        key=lambda x: x[1])[-1] + 1)],
                            subplot_titles=tuple(headers),
                            horizontal_spacing=0.00,
                            vertical_spacing=0.03)

        # add traces
        for idx, t in enumerate(traces):
            fig.add_trace(t, col=idx_to_grid[idx][0] + 1, row=idx_to_grid[idx][1] + 1)

        # layout (title, margins)
        fig.update_layout(title={"text": title, "x": 0.5, "xanchor": "center"},
                          margin=dict(l=0, r=0, b=0))

        # update subtitle sizes
        # fig.for_each_annotation(lambda a: a.update(text=f"<b>{a.text}</b>"))
        fig.update_annotations(font_size=10)
        return fig


class MeSHSunburst(SunburstBase):
    """Phenotype/MeSH Sunburst Class"""

    def __init__(self):
        super().__init__()
        self.database = None
        self.is_init = False

        self.drug_name = None
        self.phenotype_counts = dict()
        self.mesh_tree = dict()
        self.mesh_to_tree_id = dict()  # 1:N mesh to mesh-tree-ids
        self.plot_type = {
            "Sunburst Plot": Sunburst,
            "Icicle Plot": Icicle,
            "Treemap": Treemap
        }

    def init(self, database: str = None) -> None:
        """Manual database initialization routine

        :param database: path to database
        """
        if database:
            self.database = database
            super().init_db(self.database)
            self.init_mesh_tree()

        self.is_init = True

    def init_mesh_tree(self) -> None:
        """Initializes and loads MeSH-tree without counts and default color into memory"""
        # load base mesh tree
        self.mesh_tree = {k[0]: {} for k in self.query("SELECT id FROM mesh_tree WHERE level = 0")}
        for line in self.query("SELECT * FROM mesh_tree"):
            _id, _name, _description, _mesh_id, _parent, _level = line
            self.mesh_tree[_id.split(".")[0]][_id] = {
                "id": _id,
                "label": _name,
                "description": _description,
                "comment": "",
                "mesh_id": _mesh_id,
                "parent": _parent,
                "level": _level,
                "counts": 0,
                "color": self.s["default_color"]
            }
        self.populate_mesh_to_tree_id()

        print(f"Loaded MeSH-tree with {len(self.mesh_tree)} main nodes into memory")

    def populate_mesh_to_tree_id(self) -> None:
        """Populate mesh_to_tree_id lookup dict used for exporting MeSH ontology"""
        self.mesh_to_tree_id = dict()
        for main_id, node in self.mesh_tree.items():
            for node_id, node_data in node.items():
                mesh_id = node_data["mesh_id"]
                if mesh_id not in self.mesh_to_tree_id.keys():
                    self.mesh_to_tree_id[mesh_id] = set()
                self.mesh_to_tree_id[mesh_id].add(node_id)

    def export_mesh_tree(self, mode: str = "Excel", template: bool = False,
                         current_data: list = None) -> str:
        """Export mesh tree as Excel/TSV file; Primary identifier is the MeSH ID

        :param mode: defines whether to create an Excel or .tsv file
        :param template: if True, a template is created (all-white, 0 counts)
        :param current_data: use custom data and export plot as it is currently displayed in UI
        :returns: absolute path to generated Excel file (filename: mesh_tree_{drug_name}.xlsx/tsv
        """
        print("Exporting MeSH-tree ..")
        if template:
            fn_base = "mesh_tree_template"
            header = ["MeSH ID", "Tree ID", "Name", "Description", "Comment",
                      "Counts [Template Drug]", "Color"]
        else:
            if self.custom_ontology:
                fn_base = sanitize_string(self.custom_ontology_title)
                header = ["ID", "Parent", "Label", "Description", "Count", "Color"]
            else:
                fn_base = f"mesh_tree_{self.drug_name.lower()}"
                header = ["MeSH ID", "Tree ID", "Name", "Description", "Comment",
                          f"Counts [{self.drug_name}]", "Color"]
                self.populate_mesh_to_tree_id()

        if current_data:
            current_data = self.get_label_to_current_counts(current_data)

        is_a_exists = self.verify_is_a_attribute_exists()

        # get unique rows based on MeSH-id
        unique_rows = set()
        dupe_check = set()
        white = "#FFFFFF"
        for sub_tree_id, sub_tree in self.mesh_tree.items():
            for node_id, node in sub_tree.items():
                if self.custom_ontology:
                    node_id = node["id"]
                else:
                    node_id = node["mesh_id"]

                # if not custom ontology, skip dupes (have same counts, colors anyway)
                if not self.custom_ontology and node_id in dupe_check:
                    continue

                # add row data
                label = node["label"]
                description = node["description"]
                if self.custom_ontology:
                    parent = node["parent"]
                    if parent and is_a_exists and len(node["is_a"]) > 1:
                        parent = "|".join([_[0] for _ in node["is_a"]])

                    # minimal format with 4 columns
                    unique_rows.add((node_id,
                                     parent,
                                     label,
                                     description.replace("\n", ";"),
                                     0,
                                     white))
                else:
                    if current_data:
                        # replaces counts with propagated counts
                        counts = current_data[label]
                    else:
                        counts = int(node["counts"])
                    unique_rows.add((node_id,
                                     "|".join(self.mesh_to_tree_id[node_id]),
                                     label,
                                     description,
                                     "",
                                     counts if not template else 0,
                                     node["color"] if not template else white))

                # add mesh id to dupe check
                dupe_check.add(node_id)

        # sort by counts
        if self.custom_ontology:
            unique_rows = sorted(unique_rows, key=lambda x: x[4], reverse=True)
        else:
            unique_rows = sorted(unique_rows, key=lambda x: x[5], reverse=True)

        if mode == "Excel":
            # get general & mesh-related settings
            settings = [(k, v) for k, v in self.s.items()
                        if not k.startswith("atc_") and k != "default_color"]

            # write to .xlsx file, return filename
            return self.export_tree_to_excel(fn_base + ".xlsx", header, unique_rows, settings, 7)

        elif mode == "TSV":
            # write to .tsv file, return filename
            return self.export_tree_to_tsv(fn_base + ".tsv", header, unique_rows)

    def verify_is_a_attribute_exists(self) -> bool:
        for sub_tree_id, sub_tree in self.mesh_tree.items():
            for node_id, node in sub_tree.items():
                if "is_a" in node.keys():
                    return True
                else:
                    return False

    def read_mesh_settings_from_excel(self, wb: Workbook = None, fn: str = None) -> None:
        """Read settings from excel and apply to core object

        :param wb: Workbook object
        :param fn: Excel filename if no Workbook object was given
        """
        if not wb:
            wb = load_workbook(fn, read_only=True)

        ws_settings = wb["Settings"]
        settings = {r[0].value: r[1].value for r in ws_settings.rows}
        self.set_settings(settings)

    def _reconstruct_separator_based_tree(self, tree_ids: str = None,
                                          level_separator: str = None,
                                          id_separator: str = "|", **kwargs) -> None:
        """Process tree ids, reconstruct mesh tree

        :param tree_ids: single entity ID or list of entities separated with id_separator
        :param level_separator: separator between levels, e.g. "." for "C01.001" for MeSH
        :param id_separator: separator between ids, e.g. "|" for "C01.001|C01.002"
        """
        # process tree ids, reconstruct mesh tree
        for tree_id in tree_ids.split(id_separator):
            main_id = tree_id.split(level_separator)[0]
            level = tree_id.count(level_separator)
            parent = tree_id.rsplit(level_separator, 1)[0] if level > 0 else ""
            if main_id not in self.mesh_tree.keys():
                self.mesh_tree[main_id] = {}
            self.mesh_tree[main_id][tree_id] = {
                # "counts": counts,
                # "label": name,
                # "description": description,
                # "comment": comment,
                # "color": color,
                "id": tree_id,
                "level": level,
                "parent": parent,
                # "mesh_id": mesh_id
                **kwargs
            }

            # validate all parents exist
            self.check_mesh_parent(parent=parent, main_id=main_id, separator=level_separator)

    def check_mesh_parent(self, parent: str = None, main_id: str = None,
                          separator: str = None) -> None:
        """Creates artificial parent node if not existent > checks parent's parent availability"""
        if parent and parent not in self.mesh_tree[main_id].keys():
            parents_parent = parent.rsplit(separator, 1)[0]
            level = parent.count(separator)
            self.mesh_tree[main_id][parent] = {
                "counts": self.zero,
                "label": "N/A",
                "description": "Undefined",
                "comment": "",
                "color": self.s["default_color"],
                "id": parent,
                "level": level,
                "parent": parents_parent if level > 0 else "",
                "mesh_id": ""
            }

            # check next parents existance
            self.check_mesh_parent(parent=parents_parent, main_id=main_id, separator=separator)

    def process_custom_row_data(self, row_data: [io.TextIOWrapper, object],
                                ontology_type: str = None) -> None:
        """Process a .tsv file row by row for populating custom ontologies

        :param row_data: file IO wrapper
        :param ontology_type: type of ontology
        """
        separator = "."
        if ontology_type == "custom_sep_slash":
            separator = "/"
        elif ontology_type == "custom_sep_colon":
            separator = ","
        elif ontology_type == "custom_sep_underscore":
            separator = "_"

        for idx, row in enumerate(row_data):
            if idx == 0:
                self.drug_name = "CUSTOM"
                continue

            custom_id, label, description, counts, color, *unwanted = row.rstrip("\n").split("\t")

            if not custom_id or custom_id == "":
                continue

            label, description, counts, color = self._set_default_row_data(
                custom_id, label, description, counts, color)

            self._reconstruct_separator_based_tree(
                custom_id, level_separator=separator, counts=counts, label=label,
                description=description, color=color)

            self.phenotype_counts[label] = counts

    def process_mesh_row_data(self, row_data: [io.TextIOWrapper, object]) -> None:
        """Process a .tsv or Excel file row by row

        :param row_data: either rows of a Worksheet (e.g. wb["Tree"].rows) or a file IO wrapper
        """
        for idx, row in enumerate(row_data):

            # get drug name, skip header
            if idx == 0:
                drug_name = [_.value for _ in row][-2] \
                    if isinstance(row, tuple) else row.rstrip("\n").split("\t")[-2]
                if "Counts [" in drug_name:
                    drug_name = drug_name.split("Counts [")[-1].rstrip("]")
                self.drug_name = drug_name
                continue

            # worksheet iterators return tuples and require retrieval of cell-values with cell.value
            if isinstance(row, tuple):
                (mesh_id, tree_ids, name, description, comment, counts,
                 color) = [_.value for _ in row]
            else:
                (mesh_id, tree_ids, name, description, comment, counts,
                 color) = row.rstrip("\n").split("\t")

            # skip rows without mesh id
            if not mesh_id or mesh_id == "":
                continue

            # set defaults if cells are empty, assign self.zero to cells without values
            name, description, counts, color = self._set_default_row_data(mesh_id, name,
                                                                          description, counts,
                                                                          color)

            if not comment:
                comment = ""

            # process tree ids, reconstruct mesh tree
            self._reconstruct_separator_based_tree(
                tree_ids, level_separator=".", counts=counts, label=name, description=description,
                comment=comment, color=color, mesh_id=mesh_id)

            # update phenotype counts
            self.phenotype_counts[name] = counts

        print(f"\tAdded {self.get_total_counts(count_key='counts')} "
              f"counts for drug '{self.drug_name}'")

    def populate_mesh_from_tsv(self, fn: str = None, **kwargs) -> None:
        """Populate MeSH tree from tsv data

        :param fn: path to .tsv file
        """
        self.rollback_mesh_tree()
        print(f"Loading MeSH-tree from {fn} ..")
        with open(fn, mode="r", encoding="utf-8") as f_in:
            self.process_mesh_row_data(row_data=f_in)

    def populate_custom_ontology_from_tsv(self, fn: str = None, ontology_type: str = None) -> None:
        """Populates a custom ontology from tsv data

        :param fn: path to .tsv file
        :param ontology_type: type of ontology to parse
        """
        self.rollback_mesh_tree()
        print(f"Loading data from {fn} ..")
        with open(fn, mode="r", encoding="utf-8") as custom_file:
            if ontology_type.startswith("custom_sep_"):
                self.process_custom_row_data(row_data=custom_file, ontology_type=ontology_type)

        # set all zero counts to fake_one to force display of all sub-trees
        for sub_tree in self.mesh_tree.values():
            for node in sub_tree.values():
                if node["counts"] == self.zero:
                    node["imported_counts"] = self.fake_one
                else:
                    node["imported_counts"] = node["counts"]

    def populate_custom_ontology_from_web(self) -> None:
        """Copies already populated tree based on streamed .obo file, populates phenotype_counts"""
        self.rollback_mesh_tree()
        self.mesh_tree = self.custom_ontology
        for sub_tree in self.mesh_tree.values():
            for node in sub_tree.values():
                self.phenotype_counts[node["label"]] = node["counts"]

    def load_mesh_excel(self, fn: [str, None] = None, read_settings: bool = True,
                        populate: bool = True) -> None:
        """Load data from MeSH-specific Excel file; Allows new nodes to be added, color/count modifications;

         :param fn: path to Excel file containing MeSH-id, counts, color, ..
         :param read_settings: If True, settings from Excel will be loaded and applied
         :param populate: If True, MeSH tree is loaded and processed
        """
        wb = load_workbook(fn, read_only=True)
        self.rollback_mesh_tree()

        # read & iterate over excel - load settings
        if read_settings:
            self.read_mesh_settings_from_excel(wb=wb)

        # load tree data
        if populate:
            print(f"Loading MeSH-tree from {fn} ..")
            try:
                ws = wb["Tree"]
            except KeyError:
                ws = wb.worksheets[0]
            self.process_mesh_row_data(ws.rows)

        wb.close()

    def rollback_mesh_tree(self, hard_reset: bool = True) -> None:
        """Clears counts and resets color of mesh-tree"""
        if hard_reset:
            self.mesh_tree = dict()
        else:
            for main_id, node in self.mesh_tree.items():
                for sub_node, v in node.items():
                    v["counts"] = 0
                    v["imported_counts"] = 0
                    v["color"] = self.s["default_color"]
                    v["comment"] = ""
        self.phenotype_counts = dict()
        self.drug_name = None

    def populate_mesh_from_data_source(
            self, drug_name: str = None,
            data_source: str = "Utilization Tuple: Semantic Direct") -> None:
        """Populates the MeSH tree from a data source (database)

        :param drug_name: If string with drug-name is given, fetch phenotype counts
            and repopulate tree
        :param data_source: Data source as string (handed over from GUI)
            Possible values:
                "Utilization Tuple: Semantic Direct"
                "Utilization Tuple: Semantic Indirect"
                "Utilization Tuple: Explicit Direct"
                "Utilization Tuple: Explicit Indirect"
            Explanation:
                "semantic" considers text-mined associations, while "explicit" is based
                    on expert-rated annotation
                "indirect" includes known targets and markers for selected drug <> phenotype pairs
        """
        print(f"Populating MeSH tree from data source: {data_source} ..")
        self.rollback_mesh_tree()

        # fetch drug id
        drug_id = self.get_drug_id(drug_name)

        # resolve data source
        qry = None
        if data_source in ["Utilization Tuple: Semantic Direct",
                           "Utilization Tuple: Semantic Indirect",
                           "Utilization Tuple: Explicit Direct",
                           "Utilization Tuple: Explicit Indirect"]:
            if "Semantic" in data_source:
                semantic = True
            else:
                semantic = False

            if "Indirect" in data_source:
                indirect = True
            else:
                indirect = False

            target_db = str("pheno_" + ["direct", "indirect"][indirect] + "_"
                            + ["explicit", "semantic"][semantic])
            qry = "SELECT phenotype_id, cnt FROM {} WHERE drug_asset=?".format(target_db)

        # fetch phenotype counts
        ret = self.query(qry, [drug_id])

        # populate class variables
        self.drug_name = drug_name
        self.phenotype_counts = {self.phenotype_lookup_reverse[_[0]]: _[1] for _ in ret}

        # populate mesh tree with recent phenotype-counts and respective colors
        for main_id, node in self.mesh_tree.items():

            # update phenotype counts for selected drug
            for child_id, v in node.items():
                v["counts"] = self.phenotype_counts.get(v["label"], 0)
                v["imported_counts"] = self.phenotype_counts.get(v["label"], 0)

            # calculate color scale based on sub trees max value
            factor, scale = self.calculate_color_scale_for_node(node)

            # apply colors
            for child_id, v in node.items():
                v["color"] = scale[int(v["counts"] / factor)]

        print(f"\tAdded {self.get_total_counts(count_key='counts')} counts for "
              f"drug '{self.drug_name}'")

    def plot(self) -> None:
        """Generate data for phenotype sunburst plot"""
        self.set_thread_status("Creating separator-based sunburst ..")
        self.thread_return = None
        self.plot_error = None
        plot_tree = {}
        parent_whitelist = set()
        drop_count = 0

        # create copy of tree
        # first level keys are sorted C01, C02
        # inner keys are sorted by level (outer to inner)
        for k, v in sorted(self.mesh_tree.items()):

            # if all values of sub-tree are zero, skip copy
            if self.s["mesh_drop_empty_last_child"] and all(_['counts'] == self.zero
                                                            for _ in v.values()):
                self.set_thread_status(f"Skipping sub-tree {k} without values")
                continue

            if k not in plot_tree.keys():
                plot_tree[k] = {}
            for kk, vv in sorted(v.items(), key=lambda x: x[1]["level"], reverse=True):
                
                # drop empty nodes
                counts = vv["counts"]
                if self.s["mesh_drop_empty_last_child"] and counts == self.zero and vv["id"] \
                        not in parent_whitelist:
                    drop_count += 1
                    continue

                # add childs parent id to parent_whitelist to not remove empty parents
                parent_whitelist.add(vv["parent"])

                # copy node, set counts to at least self.fake_one
                # to ensure all nodes (0-counts) are displayed
                plot_tree[k][kk] = vv
                plot_tree[k][kk]["counts"] = vv["counts"] if vv["counts"] >= 1 else self.fake_one
                plot_tree[k][kk]["imported_counts"] = counts

        if self.s["mesh_drop_empty_last_child"]:
            self.set_thread_status(f"Dropped {drop_count} empty child nodes ..")

        # propagate counts up
        if self.s["mesh_propagate_enable"]:
            self.set_thread_status("Propagating counts ..")
            for k, v in plot_tree.items():
                for kk, vv in v.items():

                    # skip if no further parent exists
                    try:
                        parent = plot_tree[k][vv["parent"]]
                    except KeyError:
                        continue

                    # apply count propagation
                    propagate_mode = self.s["mesh_propagate_counts"]
                    if propagate_mode == "off":
                        continue
                    elif propagate_mode == "level":
                        if parent["level"] >= self.s["mesh_propagate_lvl"]:
                            parent["imported_counts"] += vv["imported_counts"]
                    elif propagate_mode == "all":
                        parent["imported_counts"] += vv["imported_counts"]

        # when counts are propagated, begin color propagation
        self.tree_color_propagation(plot_tree=plot_tree, count_key="imported_counts")

        # create figure
        try:
            self.create_sunburst_figure(plot_tree=plot_tree)
        except Exception as exc:
            self.plot_error = exc


class ATCSunburst(SunburstBase):
    """Drug/ATC Phenotype Class"""

    def __init__(self):
        super().__init__()
        self.database = None
        self.is_init = False

        self.phenotype_name = None
        self.drug_counts = dict()
        self.atc_tree = dict()
        self.id_to_chembl = dict()
        self.chembl_to_id = dict()
        self.chembl_to_drug_name = dict()

    def init(self, database: str = None) -> None:
        """Manual database initialization routine

        :param database: path to database
        """
        if database:
            self.database = database
            super().init_db(self.database)
            self.init_atc_tree()

        self.is_init = True

    def init_atc_tree(self) -> None:
        """Initializes and loads ATC-tree without counts and default color into memory"""

        # populate atc_tree base-levels from level 1 codes
        self.atc_tree = {k[0]: {} for k in self.query("SELECT DISTINCT level1 FROM drug_atc")}

        # populate chembl <> drug lookup dicts; id_to_chembl is 1:N, chembl_to_id 1:1
        for row in self.query("SELECT * FROM drug_lookup"):
            _id, drug_name, chembl_id = row
            if _id not in self.id_to_chembl.keys():
                self.id_to_chembl[_id] = set()
            self.id_to_chembl[_id].add(chembl_id)
            self.chembl_to_id[chembl_id] = _id

            # populate chembl <> drug name as in platform to be consistent when exporting excel
            self.chembl_to_drug_name[chembl_id] = drug_name

        # populate atc_tree sub-trees
        for row in self.query("SELECT * FROM drug_atc WHERE chembl_id IN "
                              "(SELECT chembl_id FROM drug_lookup)"):
            chembl_id, drug_name, levels, descriptions = row[0], row[1], row[2:7], row[7:]
            level_one = levels[0]
            labels = (*descriptions, drug_name)

            for idx, lvl in enumerate(levels):
                if not self.atc_tree.get(level_one).get(lvl):
                    self.atc_tree[level_one][lvl] = {
                        "label": labels[idx],
                        "comment": "",
                        "counts": 0,
                        "imported_counts": 0,
                        "counts_corrected": False,
                        "id": lvl,
                        "parent": levels[idx-1] if idx+1 > 1 else "",
                        "level": idx+1,
                        "chembl_ids": set(),
                        "color": self.s["default_color"]
                    }
                self.atc_tree[level_one][lvl]["chembl_ids"].add(chembl_id)

        print(f"Loaded ATC-tree with {len(self.atc_tree)} main nodes into memory")

    def export_atc_tree(self, mode: str = "Excel", template: bool = False) -> str:
        """Export level 5 ATC entries to Excel; Identifier is the ATC code

        :param mode: defines whether to create an Excel or .tsv file
            (filename: atc_tree_{phenotype_name}.xlsx/tsv)
        :param template: if True, a template is created (all-white, 0 counts)
        :returns: absolute path of generated Excel file
        """
        print("Exporting ATC-tree ..")
        if template:
            fn_base = "atc_tree_template"
            header = ["ATC code", "Level", "Label", "Comment",
                      "Counts [Template Phenotype]", "Color"]
        else:
            fn_base = f"atc_tree_{self.phenotype_name.lower()}"
            header = ["ATC code", "Level", "Label", "Comment",
                      f"Counts [{self.phenotype_name}]", "Color"]

        # get unique rows based on ATC code
        unique_rows = set()
        for main_id, node in self.atc_tree.items():
            for sub_id, v in node.items():
                unique_rows.add((sub_id,
                                 int(v["level"]),
                                 v["label"],
                                 v["comment"],
                                 int(v["counts"]) if not template else 0,
                                 v["color"] if not template else "#FFFFFF"))

        # sort by level > counts
        unique_rows = sorted(unique_rows, key=lambda x: (x[1], x[4]), reverse=True)

        if mode == "Excel":
            # get general & atc-related settings
            settings = [(k, v) for k, v in self.s.items()
                        if not k.startswith("mesh_") and k != "default_color"]

            # write to file, return filename
            return self.export_tree_to_excel(fn_base + ".xlsx", header, unique_rows, settings, 6)

        elif mode == "TSV":
            # write to .tsv file, return filename
            return self.export_tree_to_tsv(fn_base + ".tsv", header, unique_rows)

    def rollback_atc_tree(self, hard_reset: bool = True) -> None:
        """Reset counts / colors of ATC tree"""
        if hard_reset:
            self.atc_tree = dict()
        else:
            for main_id, node in self.atc_tree.items():
                for sub_id, v in node.items():
                    v["counts"] = 0
                    v["imported_counts"] = 0
                    v["color"] = self.s["default_color"]
                    v["comment"] = ""
        self.drug_counts = dict()
        self.phenotype_name = None

    def clear_non_drug_counts(self) -> None:
        """Clears ATC counts for level 1-4"""
        for node in self.atc_tree.values():
            for v in node.values():
                if v["level"] != 5:
                    v["counts"] = self.zero

    def read_atc_settings_from_excel(self, wb: Workbook = None, fn: str = None) -> None:
        """Read settings from excel and apply to core object

        :param wb: Workbook object
        :param fn: Excel filename if no workbook is given
        """
        if not wb:
            wb = load_workbook(fn, read_only=True)

        ws_settings = wb["Settings"]
        settings = {r[0].value: r[1].value for r in ws_settings.rows}

        if settings["atc_propagate_to_level"] != -1:
            popup = Tk()
            popup.withdraw()
            messagebox.showwarning("Propagation warning",
                                   "WARNING - propagation active - custom colors will be "
                                   "overwritten! Set 'atc_propagate_to_level' to '1' to "
                                   "enable display of custom colors")
            popup.destroy()
            print("WARNING - propagation active - custom colors will be overwritten! "
                  "Set atc_propagate_to_level to 1 to prevent")

        self.set_settings(settings)

    def populate_atc_from_tsv(self, fn: str = None, **kwargs) -> None:
        """Populate ATC tree from tsv data

        :param fn: path to .tsv file
        """
        self.rollback_atc_tree()
        print(f"Loading ATC-tree from {fn} ..")
        with open(fn, mode="r", encoding="utf-8") as f:
            self.process_atc_row_data(f)

    def check_atc_parent(self, parent: str, tree_id: str, parents_level: int) -> None:
        """Creates artificial parent node if not existent > checks parent's parent availability"""
        if parent and parent != "" and parent not in self.atc_tree[tree_id].keys():
            parents_parent = parent[:-2] if parents_level in [5, 2] else parent[:-1]
            self.atc_tree[tree_id][parent] = {
                "label": "",
                "counts": self.zero,
                "comment": "",
                "imported_counts": self.zero,
                "counts_corrected": False,
                "id": parent,
                "parent": parents_parent,
                "level": parents_level,
                "color": self.s["default_color"]
            }

            # check next parents existence
            self.check_atc_parent(parents_parent, tree_id, parents_level-1)

    def process_atc_row_data(self, row_data: [io.TextIOWrapper, object]) -> None:
        """Process a .tsv or Excel file row by row

        row_data: either rows of a Worksheet (e.g. wb["Tree"].rows) or a file IO wrapper
        """
        for idx, row in enumerate(row_data):

            # get phenotype name, skip header
            if idx == 0:
                pheno_name = [_.value for _ in row][-2] \
                    if isinstance(row, tuple) else row.rstrip("\n").split("\t")[-2]
                if "Counts [" in pheno_name:
                    pheno_name = pheno_name.split("Counts [")[-1].rstrip("]")
                self.phenotype_name = pheno_name
                continue

            # worksheet iterators return tuples and require retrieval of cell-values with cell.value
            if isinstance(row, tuple):
                atc_code, level, label, comment, counts, color = [_.value for _ in row]
            else:
                atc_code, level, label, comment, counts, color = row.rstrip("\n").split("\t")

            # skip rows without atc code or level
            if not atc_code or not level or atc_code == "" or level == "":
                continue

            # set defaults if cells are empty, assign self.zero to cells without values
            label, comment, counts, color = self._set_default_row_data(atc_code, label, comment,
                                                                       counts, color)

            if isinstance(level, str):
                level = int(level)

            if not comment:
                comment = ""

            # process atc code, reconstruct atc tree
            parent = ""
            if level in [5, 2]:
                parent = atc_code[:-2]
            elif level in [4, 3]:
                parent = atc_code[:-1]

            if not atc_code[0] in self.atc_tree.keys():
                self.atc_tree[atc_code[0]] = {}
            self.atc_tree[atc_code[0]][atc_code] = {
                "label": label,
                "counts": counts,
                "comment": comment,
                "imported_counts": counts,
                "counts_corrected": False,
                "id": atc_code,
                "parent": parent,
                "level": level,
                "color": color
            }

            # validate all parents exist
            self.check_atc_parent(parent=parent, tree_id=atc_code[0], parents_level=level-1)

        # validate parent counts sum up to child counts while ignoring color
        # difference may be introduced by adding customized counts
        # meaning: counts for atc codes > level 5 will be overwritten to enable wedge-width 'total'
        self.clear_non_drug_counts()
        print(f"\tAdded {self.get_total_counts(count_key='counts')} counts "
              f"for phenotype '{self.phenotype_name}'")

    def load_atc_excel(self, fn: str = None, read_settings: bool = True,
                       populate: bool = True) -> None:
        """Load data from ATC-specific Excel file

        :param fn: path to Excel file
        :param read_settings: If True, settings from Excel will be loaded and applied
        :param populate: If True, ATC tree is loaded and processed
        """
        work_book = load_workbook(fn, read_only=True)
        self.rollback_atc_tree()

        # read & iterate over excel - load settings
        if read_settings:
            self.read_atc_settings_from_excel(wb=work_book)

        # load tree data
        if populate:
            print(f"Loading ATC-tree from {fn} ..")
            try:
                work_sheet = work_book["Tree"]
            except KeyError:
                work_sheet = work_book.worksheets[0]
            self.process_atc_row_data(row_data=work_sheet.rows)

        work_book.close()

    def populate_atc_from_data_source(self, phenotype_name: str = None,
                                      data_source: str = None) -> None:
        """Populates the ATC tree from a data source (database)

        :param phenotype_name: Phenotype name
        :param data_source: Data source as string (handed over from GUI); Possible values: 'Linked Tuples'
        """

        print(f"Populating ATC tree from data source: {data_source} ..")
        self.rollback_atc_tree()

        # fetch phenotype id
        phenotype_id = self.get_entity_id(phenotype_name, "phenotype")

        # resolve data source:
        qry = None
        if data_source == "Linked Tuple":
            qry = "SELECT drug_id FROM drug_lt WHERE phenotype_id=?"

        # fetch drug counts
        self.drug_counts = {}
        self.phenotype_name = phenotype_name
        for drug_id in self.query(qry, [phenotype_id]):
            for chembl_id in self.id_to_chembl[drug_id[0]]:
                if chembl_id not in self.drug_counts.keys():
                    self.drug_counts[chembl_id] = 0
                self.drug_counts[chembl_id] += 1

        # add drug counts directly based on chembl_id(s) to level 5
        for node in self.atc_tree.values():
            for val in node.values():
                if val["level"] == 5:
                    for chembl_id in val["chembl_ids"]:
                        if chembl_id in self.drug_counts.keys():
                            val["counts"] += self.drug_counts[chembl_id]
                            val["imported_counts"] += self.drug_counts[chembl_id]

            # calculate color scale, apply to level 5 only
            factor, scale = self.calculate_color_scale_for_node(node)
            for val in node.values():
                if val["level"] == 5:
                    val["color"] = scale[int(val["counts"] / factor)]

        print(f"\tAdded {self.get_total_counts(count_key='counts')} counts "
              f"for phenotype '{self.phenotype_name}'")

    def plot(self):
        """Generate data for drug sunburst plot"""
        self.set_thread_status("Creating drug sunburst ..")
        self.thread_return = None
        self.plot_error = None

        # create & sort plot tree
        plot_tree = dict(sorted(self.atc_tree.items()))

        # setup counts, propagate if enabled
        for key, val in plot_tree.items():
            for inner_val in val.values():

                # set all level 5 nodes to at least self.fake_one if loaded from file
                if inner_val["level"] == 5:
                    if inner_val["imported_counts"] <= 1:
                        inner_val["counts"] = self.fake_one

                # reset all other levels counts to 0
                else:
                    inner_val["counts"] = 0

            # propagate counts up from level 5 > 1
            for inner_key, inner_val in sorted(val.items(), key=lambda x: x[1]["level"],
                                               reverse=True):
                if inner_val["parent"] != "":
                    plot_tree[key][inner_val["parent"]]["counts"] += inner_val["counts"]

                    # propagate counts (overwrite imported counts) if enabled
                    propagate_mode = self.s["atc_propagate_counts"]
                    if propagate_mode == "level":
                        if inner_val["level"] > self.s["atc_propagate_lvl"]:
                            plot_tree[key][inner_val["parent"]]["imported_counts"] += inner_val[
                                "imported_counts"]
                    elif propagate_mode == "all":
                        plot_tree[key][inner_val["parent"]]["imported_counts"] += inner_val[
                            "imported_counts"]

        # when counts are propagated, begin color propagation
        self.tree_color_propagation(plot_tree=plot_tree, count_key="imported_counts")

        # create figure
        try:
            self.create_sunburst_figure(plot_tree=plot_tree)
        except Exception as exc:
            self.plot_error = exc


def show_help():
    """prints help to console"""
    print(" DrugVision - Minimal Reproducible Examples ".center(120, "-"))
    print('''
from ontoloviz.core import MeSHSunburst, ATCSunburst

""" phenotype sunburst with connected database """
p = MeSHSunburst()
p.init("drugvision.db")
p.populate(drug_name="Aspirin")
p.export_mesh_tree()
p.plot()

""" phenotype sunburst from Excel file """
p = MeSHSunburst()
p.populate_mesh_from_excel("mesh_tree_aspirin.xlsx")
p.plot()

""" drug sunburst with connected database """
d = ATCSunburst()
d.init(db)
d.populate(phenotype_name="headache")
d.export_atc_tree()
d.plot()

""" drug sunburst from Excel file """
d = ATCSunburst()
d.populate_atc_from_excel("atc_tree_headache.xlsx")
d.plot
''')


if __name__ == "__main__":
    show_help()
